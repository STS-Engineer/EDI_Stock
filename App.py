from flask import Flask, request, render_template_string, send_file, Response, abort
import pandas as pd
import os
import time
import chardet
import io
import math
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import base64, pdfplumber, re
from dateutil.parser import parse as dtparse
app = Flask(__name__)

# -----------------------
# Database Connection
# -----------------------
# Ensure you replace these with your actual database credentials
db_url = URL.create(
    drivername="postgresql+psycopg2",
    username="administrationSTS",
    password="St$@0987",
    host="avo-adb-002.postgres.database.azure.com",
    port=5432,
    database="EDI_IA"
)
try:
    engine = create_engine(db_url)
    # Test connection
    with engine.connect() as connection:
        print("✅ Database connection successful.")
except Exception as e:
    print(f"❌ Database connection failed: {e}")
    engine = None

OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)
ALLOWED_EXTENSIONS = {"csv", "xls", "xlsx", "pdf"}


import re
from math import isnan

def _clean_qty(v):
    """
    Convert various user-entered quantity formats to a safe int.
    Handles: strings with spaces/commas, floats, NaN, None.
    """
    if v is None:
        return 0
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        # If it's NaN, treat as 0
        return 0 if (v != v) else int(round(v))
    s = str(v).strip()
    if s == "" or s.lower() == "nan" or s.lower() == "none":
        return 0
    # remove common thousand separators and non-digit except minus sign
    s = s.replace(",", "").replace(" ", "").replace("\u00A0", "")  # remove normal & non-breaking spaces
    m = re.match(r"^-?\d+(\.\d+)?$", s)
    if m:
        return int(float(s))
    # last resort: strip everything not a digit or minus
    s2 = re.sub(r"[^\d-]", "", s)
    return int(s2) if s2 not in ("", "-",) else 0

def _norm_status(s):
    if s is None:
        return ""
    s = str(s).strip()
    if s.lower() == "sent":
        return "Dispatched"
    # unify common variants
    if s.lower().replace(" ", "") in ("intransit","in-transit"):
        return "InTransit"
    if s.lower() == "dispatched":
        return "Dispatched"
    if s.lower() == "delivered":
        return "Delivered"
    return s  # fallback unchanged




def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# -----------------------
# DB Insert Functions
# -----------------------
def insert_ediglobal(df):
    if not engine:
        raise ConnectionError("Database engine is not available.")
    with engine.begin() as conn:
        for _, row in df.iterrows():
            stmt = text("""
                INSERT INTO "EDIGlobal" (
                    "Site","ClientCode","ClientMaterialNo","AVOMaterialNo",
                    "DateFrom","DateUntil","Quantity","ForecastDate",
                    "LastDeliveryDate","LastDeliveredQuantity",
                    "CumulatedQuantity","EDIStatus","ProductName","LastDeliveryNo"
                )
                VALUES (
                    :Site,:ClientCode,:ClientMaterialNo,:AVOMaterialNo,
                    :DateFrom,:DateUntil,:Quantity,:ForecastDate,
                    :LastDeliveryDate,:LastDeliveredQuantity,
                    :CumulatedQuantity,:EDIStatus,:ProductName,:LastDeliveryNo
                )
            """)
            conn.execute(stmt, {col: row.get(col) for col in [
                "Site","ClientCode","ClientMaterialNo","AVOMaterialNo",
                "DateFrom","DateUntil","Quantity","ForecastDate",
                "LastDeliveryDate","LastDeliveredQuantity",
                "CumulatedQuantity","EDIStatus","ProductName","LastDeliveryNo"
            ]})

def insert_deliverydetails(df):
    with engine.begin() as conn:
        for _, row in df.iterrows():
            site = _safestr(row.get("Site"))
            avo_mat = _safestr(row.get("AVOMaterialNo"))
            delivery_no = _safestr(row.get("DeliveryNo"))
            date = _safestr(row.get("Date"))


            qty = _clean_qty(row.get("Quantity"))
            status = _norm_status(row.get("Status"))

            if not delivery_no or not date or not status or not site:
                continue

            def fetch_latest_intransit():
                q = text("""
                    SELECT "DeliveryNo","Quantity","Date"
                    FROM "DeliveryDetails"
                    WHERE "Site" = :site
                      AND COALESCE("AVOMaterialNo",'') = :avo_mat
                      AND "Status" = 'InTransit'
                    ORDER BY "Date" DESC
                    LIMIT 1
                """)
                return conn.execute(q, {
                    "site": site,  "avo_mat": avo_mat
                }).mappings().fetchone()

            if status == "Dispatched":
                it = fetch_latest_intransit()
                if it:
                    new_qty = _clean_qty(it["Quantity"]) + qty
                    # IMPORTANT CHANGE: also update DeliveryNo to the CURRENT dispatched delivery_no
                    update_stmt = text("""
                        UPDATE "DeliveryDetails"
                        SET "Quantity" = :qty,
                            "Date" = :date,
                            "DeliveryNo" = :curr_del_no
                        WHERE "DeliveryNo" = :old_del_no
                          AND "Status" = 'InTransit'
                          AND "Date" = :old_date
                    """)
                    conn.execute(update_stmt, {
                        "qty": int(new_qty),
                        "date": date,
                        "curr_del_no": delivery_no,
                        "old_del_no": it["DeliveryNo"],
                        "old_date": it["Date"]
                    })
                else:
                    # If your PK lets you reuse DeliveryNo here, you can set del_no=delivery_no instead of synthetic.
                    transit_no = (delivery_no + "_T")[:30]
                    insert_it = text("""
                        INSERT INTO "DeliveryDetails"
                        ("Site","AVOMaterialNo","DeliveryNo","Quantity","Date","Status")
                        VALUES (:site,:avo_mat,:del_no,:qty,:date,'InTransit')
                    """)
                    conn.execute(insert_it, {
                        "site": site,  "avo_mat": avo_mat,
                        "del_no": transit_no, "qty": int(qty), "date": date
                    })

                ins_disp = text("""
                    INSERT INTO "DeliveryDetails"
                    ("Site","AVOMaterialNo","DeliveryNo","Quantity","Date","Status")
                    VALUES (:site,:avo_mat,:del_no,:qty,:date,'Dispatched')
                """)
                conn.execute(ins_disp, {
                    "site": site,  "avo_mat": avo_mat,
                    "del_no": delivery_no, "qty": int(qty), "date": date
                })

            elif status == "Delivered":
                it = fetch_latest_intransit()
                if it:
                    new_qty = max(0, _clean_qty(it["Quantity"]) - qty)
                    # IMPORTANT CHANGE: also refresh DeliveryNo to the CURRENT delivered delivery_no
                    update_stmt = text("""
                        UPDATE "DeliveryDetails"
                        SET "Quantity" = :qty,
                            "Date" = :date,
                            "DeliveryNo" = :curr_del_no
                        WHERE "DeliveryNo" = :old_del_no
                          AND "Status" = 'InTransit'
                          AND "Date" = :old_date
                    """)
                    conn.execute(update_stmt, {
                        "qty": int(new_qty),
                        "date": date,
                        "curr_del_no": delivery_no,
                        "old_del_no": it["DeliveryNo"],
                        "old_date": it["Date"]
                    })

                ins_del = text("""
                    INSERT INTO "DeliveryDetails"
                    ("Site","AVOMaterialNo","DeliveryNo","Quantity","Date","Status")
                    VALUES (:site,:avo_mat,:del_no,:qty,:date,'Delivered')
                """)
                conn.execute(ins_del, {
                    "site": site, "avo_mat": avo_mat,
                    "del_no": delivery_no, "qty": int(qty), "date": date
                })

            elif status == "InTransit":
                ins_it = text("""
                    INSERT INTO "DeliveryDetails"
                    ("Site","AVOMaterialNo","DeliveryNo","Quantity","Date","Status")
                    VALUES (:site,:avo_mat,:del_no,:qty,:date,'InTransit')
                """)
                conn.execute(ins_it, {
                    "site": site, "avo_mat": avo_mat,
                    "del_no": delivery_no, "qty": int(qty), "date": date
                })

            else:
                ins_generic = text("""
                    INSERT INTO "DeliveryDetails"
                    ("Site","AVOMaterialNo","DeliveryNo","Quantity","Date","Status")
                    VALUES (:site,:avo_mat,:del_no,:qty,:date,:status)
                """)
                conn.execute(ins_generic, {
                    "site": site, "avo_mat": avo_mat,
                    "del_no": delivery_no, "qty": int(qty), "date": date, "status": status
                })

# -----------------------------------
# Template Generation Setup
# -----------------------------------
TEMPLATE_SCHEMAS = {
    "edi_template": [
        "Site","ClientCode","ClientMaterialNo","AVOMaterialNo",
        "DateFrom","DateUntil","Quantity","ForecastDate",
        "LastDeliveryDate","LastDeliveredQuantity",
        "CumulatedQuantity","EDIStatus","ProductName","LastDeliveryNo"
    ],
    "delivery_template": [
        "Site","AVOMaterialNo","DeliveryNo","Quantity","Date","Status"
    ],
}

EDI_NOTES = {
    "Site": "Site Avo carbon", "ClientCode": "Customer code at Avo", "ClientMaterialNo": "Part number at the customer",
    "AVOMaterialNo": "Part number at AVO", "DateFrom": "Delivery date in week format YYYY-WXX", "DateUntil": "Delivery date written as it is in the EDI File",
    "Quantity": "Requested quantity", "ForecastDate": "Write the forecast week in this format: YYYY-WXX, for example: 2025-W28",
    "LastDeliveryDate": "The most recent date on which goods were delivered to the customer in this format YYYY-WXX",
    "LastDeliveredQuantity": "The amount of goods supplied in the most recent delivery",
    "CumulatedQuantity": "The cumulative amount of goods delivered to and acknowledged by the customer over time",
    "EDIStatus": "Forcast / Firm / PO", "ProductName": "Product description", "LastDeliveryNo": "File number of the last delivery note",
}

DELIVERY_NOTES = {
    "Site": "Site Avo carbon", "AVOMaterialNo": "Part number at AVO",
    "DeliveryNo": "Delivery note number", "Quantity": "Delivered quantity", "Date": "Delivery date (YYYY-MM-DD)",
    "Status": "Delivery status (e.g., Delivered / Dispatched / In transit)",
}

def _build_excel_with_notes(headers, notes_map, n_rows=200):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        note_text = notes_map.get(header, "Enter the value for this field.")
        c.comment = Comment(note_text, "System")
        ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max(15, len(header) + 5))

    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(headers, start=1):
        prompt = notes_map.get(header, "Enter the value for this field.")
        dv = DataValidation(type="custom", formula1="TRUE", allow_blank=True, showInputMessage=True, promptTitle=header, prompt=prompt[:255])
        ws.add_data_validation(dv)
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{n_rows}")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
# -----------------------
# Responsive HTML Template
# -----------------------
HTML_PAGE = """
<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
    <title>Delivery & EDI Management</title>
    <link rel="icon" href="/static/avo_carbon.jpg" type="image/x-icon">
    <style>
      body{font-family:'Arial',sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);margin:0;padding:20px 0;min-height:100vh}
      .container{background-color:#fff;padding:30px;border-radius:15px;box-shadow:0 15px 35px rgba(0,0,0,.1);text-align:center;width:90%;max-width:1200px;margin:0 auto;overflow:hidden;backdrop-filter:blur(10px);border:1px solid rgba(255,255,255,.2)}
      .nav-tabs{display:flex;border-bottom:none;margin-bottom:30px;justify-content:center;flex-wrap:wrap;gap:10px;background:#f8f9fa;border-radius:12px;padding:8px}
      .nav-tab{background:linear-gradient(145deg,#fff,#f0f0f0);border:none;padding:15px 25px;cursor:pointer;font-size:16px;font-weight:600;text-decoration:none;color:#555;border-radius:10px;transition:all .3s ease;display:flex;align-items:center;gap:12px;min-width:200px;justify-content:center;box-shadow:0 2px 4px rgba(0,0,0,.1)}
      .nav-tab:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(0,0,0,.15);color:#333}
      .nav-tab.active{background:linear-gradient(145deg,#4f46e5,#7c3aed);color:#fff;transform:translateY(-2px);box-shadow:0 8px 25px rgba(79,70,229,.4)}
      .tab-content{display:none;animation:fadeIn .5s ease-in-out}
      .tab-content.active{display:block}
      @keyframes fadeIn{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
      h1{color:#1f2937;font-size:28px;margin-bottom:10px;font-weight:700}
      h2{color:#4f46e5;font-size:22px;margin:30px 0 20px 0;font-weight:600}
      .subtitle{font-size:16px;color:#6b7280;margin-bottom:30px;font-weight:400}
      form{display:flex;flex-direction:column;align-items:center;gap:20px;max-width:600px;margin:0 auto}
      input[type=file],input[type=text],input[type=number],select,input[type=submit],a.download-btn{padding:15px 20px;font-size:16px;border:2px solid #e5e7eb;border-radius:12px;outline:none;transition:all .3s ease;width:100%;max-width:400px;font-family:inherit;background:#fff;box-sizing:border-box}
      input[type=file]:hover,input[type=text]:hover,input[type=number]:hover,select:hover{border-color:#4f46e5;box-shadow:0 0 0 3px rgba(79,70,229,.1)}
      input[type=file]:focus,input[type=text]:focus,input[type=number]:focus,select:focus{border-color:#4f46e5;box-shadow:0 0 0 3px rgba(79,70,229,.2)}
      input[type=submit],a.download-btn{background:linear-gradient(145deg,#4f46e5,#7c3aed);color:#fff;cursor:pointer;border:none;text-decoration:none;display:inline-block;text-align:center;font-weight:600;text-transform:uppercase;letter-spacing:.5px;box-shadow:0 4px 15px rgba(79,70,229,.3)}
      input[type=submit]:hover,a.download-btn:hover{transform:translateY(-2px);box-shadow:0 8px 25px rgba(79,70,229,.4)}
      table{width:100%;border-collapse:collapse;margin-top:25px;border-radius:12px;overflow:hidden;box-shadow:0 4px 15px rgba(0,0,0,.1)}
      table th,table td{border:none;padding:15px;text-align:left}
      table th{background:linear-gradient(145deg,#4f46e5,#7c3aed);color:#fff;font-weight:600;text-transform:uppercase;letter-spacing:.5px}
      table tbody tr:nth-child(odd){background-color:#f8f9fa}
      table tbody tr:hover{background-color:#e5e7eb;transform:scale(1.01);transition:all .2s ease}
      .scrollable{overflow-x:auto;margin-top:25px;max-height:400px;border-radius:12px}
      .error-message{color:#dc2626;background:linear-gradient(145deg,#fef2f2,#fee2e2);border:2px solid #fecaca;padding:15px 20px;border-radius:12px;margin:15px auto;font-weight:500;box-shadow:0 4px 10px rgba(220,38,38,.1);max-width:600px}
      .success-message{color:#059669;background:linear-gradient(145deg,#ecfdf5,#d1fae5);border:2px solid #a7f3d0;padding:15px 20px;border-radius:12px;margin:15px auto;font-weight:500;box-shadow:0 4px 10px rgba(5,150,105,.1);max-width:600px}
      .action-group{display:flex;flex-direction:column;gap:15px;max-width:400px;margin:25px auto}
      .secondary-btn{background:linear-gradient(145deg,#059669,#047857);box-shadow:0 4px 15px rgba(5,150,105,.3)}
      .secondary-btn:hover{box-shadow:0 8px 25px rgba(5,150,105,.4)}
      footer{margin-top:50px;color:#6b7280;font-size:14px;border-top:1px solid #e5e7eb;padding-top:25px;font-weight:500}
      .logo{max-width:250px;margin-bottom:25px;filter:drop-shadow(0 4px 8px rgba(0,0,0,0.1))}
      .template-download-section{animation:slideIn .6s ease-out;margin-bottom:30px;padding:20px;background:linear-gradient(145deg,#f8f9fa,#e9ecef);border-radius:12px;border:2px solid #dee2e6}
      @keyframes slideIn{from{opacity:0;transform:translateX(-30px)}to{opacity:1;transform:translateX(0)}}
      .file-type-options{display:flex;gap:15px;justify-content:center;flex-wrap:wrap;margin:25px 0}
      @media (max-width:768px){.container{margin:10px;width:calc(100% - 20px);padding:20px}.nav-tabs{flex-direction:column}.nav-tab{min-width:auto;width:100%}}
    </style>
  </head>
  <body>
    <div class="container">
    <img src="/static/logo-avocarbon.png" alt="AvoCarbon Logo" class="logo">      
      <div class="nav-tabs">
        <a href="#" class="nav-tab" onclick="showTab('deliveries')" id="deliveries-tab">🚚 Delivery Management</a>
        <a href="#" class="nav-tab" onclick="showTab('edi')" id="edi-tab">📋 EDI Processing</a>
      </div>

      <div id="deliveries-content" class="tab-content">
        {% if deliv_msg %}
          <div class="{{ 'success-message' if deliv_ok else 'error-message' }}">{{ deliv_msg }}</div>
        {% endif %}
        
        <h1>Delivery Management System</h1>
        <p class="subtitle">Download the template, fill it with your data, and upload for processing</p>
        
        <div class="template-download-section">
          <h3 style="color: #374151;">Download Delivery Template</h3>
          <p style="color: #6b7280; font-size: 14px;">Download the template, fill it in, and upload the completed file.</p>
          <div class="file-type-options">
            <a href="/download/template/delivery_template.xlsx" class="download-btn">📥 Excel (.xlsx)</a>
          </div>
        </div>
        
        <form action="/preview" method="post" enctype="multipart/form-data">
          <input type="file" name="file" accept=".csv,.xlsx,.xls,.pdf" required>
          <input type="hidden" name="file_type" value="LIVRAISON">
          <input type="submit" value="Preview Delivery File">
        </form>

        {% if table_html and file_type == 'LIVRAISON' %}
            {% if pdf_file %}
                <h2>PDF Preview</h2>
                <div class="scrollable" style="height:600px">
                <embed src="/view/temp/{{ pdf_file }}" type="application/pdf" width="100%" height="580px"/>
                </div>
            {% endif %}
            <div class="scrollable">
                <h2>Delivery Data Preview</h2>
                <p class="subtitle">Please review the data below before processing.</p>
                {{ table_html|safe }}
            </div>
            <div class="action-group">
                <form action="/insert" method="post">
                <input type="hidden" name="temp_file" value="{{ temp_file }}">
                <input type="hidden" name="file_type" value="LIVRAISON">
                <input type="submit" value="✅ Confirm & Send to Database" class="download-btn secondary-btn">
                </form>
                <a href="/" class="download-btn" style="background: linear-gradient(145deg, #f59e0b, #d97706);">✏️ Upload New File</a>
            </div>
            {% endif %}

      </div>

      <div id="edi-content" class="tab-content">
        {% if edi_msg %}
          <div class="{{ 'success-message' if edi_ok else 'error-message' }}">{{ edi_msg }}</div>
        {% endif %}
        
        <h1>EDI Processing Center</h1>
        <p class="subtitle">Download the template, fill it with your data, and upload for processing</p>
        
        <div class="template-download-section">
          <h3 style="color: #374151;">Download EDI Template</h3>
          <p style="color: #6b7280; font-size: 14px;">Download the template, fill it in, and upload the completed file.</p>
          <div class="file-type-options">
            <a href="/download/template/edi_template.xlsx" class="download-btn">📥 Excel (.xlsx)</a>
          </div>
        </div>
        
        <form action="/preview" method="post" enctype="multipart/form-data">
            <input type="file" name="file" accept=".csv,.xlsx,.xls,.pdf" required>
            <input type="hidden" name="file_type" value="EDI">
            <input type="submit" value="Preview EDI File">
        </form>


        {% if table_html and file_type == 'EDI' %}
        <div class="scrollable">
          <h2>EDI Data Preview</h2>
          <p class="subtitle">Please review the data below before sending to database.</p>
          {{ table_html|safe }}
        </div>
        <div class="action-group">
          <form action="/insert" method="post">
            <input type="hidden" name="temp_file" value="{{ temp_file }}">
            <input type="hidden" name="file_type" value="EDI">
            <input type="submit" value="✅ Confirm & Send to Database" class="download-btn secondary-btn">
          </form>
          <a href="/" class="download-btn" style="background: linear-gradient(145deg, #f59e0b, #d97706);">✏️ Upload New File</a>
        </div>
        {% endif %}
      </div>

      <footer>
        &copy; 2025 Delivery & EDI Management System. All rights reserved. Powered by STS AI Team
      </footer>
    </div>

    <script>
      function showTab(tabName) {
        document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
        document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
        
        document.getElementById(tabName + '-content').classList.add('active');
        document.getElementById(tabName + '-tab').classList.add('active');
        
        sessionStorage.setItem('activeTab', tabName);
      }
      
      document.addEventListener('DOMContentLoaded', function() {
        const initialTab = '{{ active_tab | default("deliveries") }}';
        showTab(initialTab);
        
        document.querySelectorAll('form').forEach(form => {
            form.addEventListener('submit', function() {
                const submitBtn = form.querySelector('input[type="submit"]');
                if (submitBtn) {
                    submitBtn.value = 'Processing...';
                    submitBtn.disabled = true;
                }
            });
        });
      });
    </script>
  </body>
</html>
"""

# -----------------------
# Flask Routes
# -----------------------
@app.route("/")
def index():
    return render_template_string(HTML_PAGE, active_tab='deliveries')


@app.route("/insert", methods=["POST"])
def insert():
    temp_file = request.form.get("temp_file")
    file_type = request.form.get("file_type")
    active_tab = 'edi' if file_type == 'EDI' else 'deliveries'

    if not temp_file:
        error_msg = "Temporary file is missing. Please try again."
        return render_template_string(HTML_PAGE, active_tab=active_tab,
                                      edi_msg=error_msg, edi_ok=False, deliv_msg=error_msg, deliv_ok=False)

    path = os.path.join(OUTPUT_DIR, secure_filename(temp_file))

    if not os.path.exists(path):
        error_msg = "Temporary file not found. It may have expired. Please upload again."
        return render_template_string(HTML_PAGE, active_tab=active_tab,
                                      edi_msg=error_msg, edi_ok=False, deliv_msg=error_msg, deliv_ok=False)

    try:
        # Read everything as strings to avoid NaN/float issues.
        df = pd.read_csv(path, dtype=str, keep_default_na=False, na_filter=False)

        if file_type == "EDI":
            insert_ediglobal(df)
            success_msg = f"✅ EDI data inserted successfully: {len(df)} rows added."
            return render_template_string(HTML_PAGE, active_tab=active_tab, edi_msg=success_msg, edi_ok=True)

        elif file_type == "LIVRAISON":
            # Normalize expected columns
            for col in ["Site","AVOMaterialNo","DeliveryNo","Date","Status"]:
                if col not in df.columns:
                    df[col] = ""
                df[col] = df[col].map(_safestr)
            df["AVOMaterialNo"] = df.apply(
                    lambda r: _normalize_avo_ref(r.get("AVOMaterialNo"), None),
                    axis=1
                )
            # Quantity as clean int
            if "Quantity" not in df.columns:
                df["Quantity"] = 0
            df["Quantity"] = df["Quantity"].apply(_clean_qty).astype(int)

            # Status normalization
            df["Status"] = df["Status"].apply(_norm_status)

            # ---- NEW: pre-aggregate duplicates (same Site/AVO/Delivery/Date/Status) ----
            key_cols = ["Site","AVOMaterialNo","DeliveryNo","Date","Status"]
            pre_count = len(df)
            df = (df.groupby(key_cols, as_index=False)["Quantity"].sum())
            # Optional: drop zeros
            df = df[df["Quantity"] != 0]
            post_count = len(df)

            # Insert with sum-aware logic (your updated insert_deliverydetails with _upsert_sum_delivery)
            insert_deliverydetails(df)

            success_msg = f"✅ Delivery data inserted successfully: {post_count} rows (aggregated from {pre_count} lines)."
            return render_template_string(HTML_PAGE, active_tab=active_tab, deliv_msg=success_msg, deliv_ok=True)

        else:
            raise ValueError("Unknown file type specified.")

    except Exception as e:
        error_msg = f"Error during database insertion: {e}"
        return render_template_string(HTML_PAGE, active_tab=active_tab,
                                      edi_msg=error_msg if active_tab == "edi" else None, edi_ok=False,
                                      deliv_msg=error_msg if active_tab == "deliveries" else None, deliv_ok=False)
    finally:
        if os.path.exists(path):
            os.remove(path)  # Clean up the temp file


@app.route("/download/template/<name>.<ext>")
def download_template(name, ext):
    headers = TEMPLATE_SCHEMAS.get(name)
    if not headers:
        abort(404, description="Unknown template name")

    if ext == "csv":
        buf = io.StringIO()
        pd.DataFrame(columns=headers).to_csv(buf, index=False)
        data = buf.getvalue().encode("utf-8-sig")
        return Response(data, mimetype="text/csv", headers={"Content-Disposition": f'attachment; filename="{name}.csv"'})

    if ext == "xlsx":
        notes_map = EDI_NOTES if name == "edi_template" else DELIVERY_NOTES
        excel_bytes = _build_excel_with_notes(headers, notes_map)
        return Response(excel_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{name}.xlsx"'})

    abort(404, description="Unsupported file extension")



def parse_delivery_pdf_bytes(pdf_bytes: bytes, *, default_site: str = "Tunisia") -> pd.DataFrame:
    import io, re, pdfplumber, pandas as pd
    from dateutil.parser import parse as dtparse

    delivery_no = None
    doc_date_iso = None
    detected_site = None
    rows = []

    # helper: add a record
    def _add(ref_val: str, qty_val: int):
        rows.append({
            "Date": doc_date_iso,
            "DeliveryNo": str(delivery_no) if delivery_no else "UNKNOWN",
            "AVOMaterialNo": ref_val,
            "Quantity": int(qty_val),
            "Site": detected_site or default_site,
            "Status": "Dispatched",
        })

    # patterns
    header_no_pat   = re.compile(r"FACTURE\s*n[°o]\s*([A-Za-z0-9\-_/]+)", re.IGNORECASE)
    header_date_pat = re.compile(r"\bDate\s+(\d{1,2}/\d{1,2}/\d{4})\b", re.IGNORECASE)
    site_pat        = re.compile(r"\bAVOCARBON[^\n]*", re.IGNORECASE)
    total_row_pat   = re.compile(r"^\s*TOTAL\b", re.IGNORECASE)

    # fallback line parser:
    # ex line: "85030010 OUI V502.730 SP PPC 11TA ... 960 1,9672 0,3262 ..."
    # capture ref (V502.730) and the quantity (960) that appears BEFORE 2–4 decimal numbers
    line_pat = re.compile(
    r"^\s*\d{8}\s+(?:OUI|NON)\s+([A-Z0-9][A-Z0-9.\-]+)(?:\s+(PL|SP))?\s+.+?\s+(\d{1,9})\s+(?:\d+[.,]\d+\s+){2,4}\S+",
    re.IGNORECASE
)


    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        # -------- header (page 1, with fallbacks) --------
        if pdf.pages:
            p0_text = pdf.pages[0].extract_text() or ""
            m_no = header_no_pat.search(p0_text)
            if m_no:
                delivery_no = m_no.group(1).strip()
            m_date = header_date_pat.search(p0_text) or re.search(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", p0_text)
            if m_date:
                doc_date_iso = dtparse(m_date.group(1), dayfirst=True).date().isoformat()
            m_site = site_pat.search(p0_text)
            if m_site:
                detected_site = "Tunisia"
        # if date still missing, try PDF metadata
        if not doc_date_iso:
            try:
                meta = pdf.metadata or {}
                meta_date = meta.get("CreationDate") or meta.get("ModDate")
                if meta_date:
                    doc_date_iso = dtparse(meta_date).date().isoformat()
            except Exception:
                pass

        # -------- attempt 1: table extraction --------
        tbl_settings = dict(
            vertical_strategy="lines",
            horizontal_strategy="lines",
            intersect_tolerance=5,
            snap_tolerance=3,
            join_tolerance=3,
            text_x_tolerance=2,
            text_y_tolerance=3,
            keep_blank_chars=False,
            edge_min_length=3,
        )
        header_ref = re.compile(r"\bREFERENCE\b|\bREFERENCE\s+ARTICLE\b|\bREF\b", re.IGNORECASE)
        header_qty = re.compile(r"\bQUANTITE\b|\bQTE\b|\bQTY\b", re.IGNORECASE)
        material_pat = re.compile(r"^[A-Za-z0-9][A-Za-z0-9.\-_/]*[A-Za-z0-9]$")

        found = 0
        for page in pdf.pages:
            tables = []
            try:
                t = page.extract_table(tbl_settings)
                if t: tables.append(t)
            except Exception:
                pass
            try:
                ts = page.extract_tables(tbl_settings) or []
                tables.extend(ts)
            except Exception:
                pass

            for tbl in tables:
                if not tbl or len(tbl) < 2:
                    continue
                ref_idx = qty_idx = None
                # find headers in first rows
                for r in tbl[:3]:
                    if not r: continue
                    for i, c in enumerate(r):
                        cell = (c or "").strip()
                        if ref_idx is None and header_ref.search(cell or ""):
                            ref_idx = i
                        if qty_idx is None and header_qty.search(cell or ""):
                            qty_idx = i
                    if ref_idx is not None and qty_idx is not None:
                        break
                # fallback heuristic on row 0
                if ref_idx is None or qty_idx is None:
                    r0 = tbl[0]
                    for i, c in enumerate(r0):
                        low = (c or "").lower()
                        if ref_idx is None and "ref" in low and "prix" not in low:
                            ref_idx = i
                        if qty_idx is None and any(k in low for k in ("quant", "qty", "qte")):
                            qty_idx = i
                if ref_idx is None or qty_idx is None:
                    continue

                data_rows = tbl[1:]
                for r in data_rows:
                    if not r: continue
                    if total_row_pat.search(" ".join([(c or "").strip() for c in r if c])):
                        continue
                    raw_ref = (r[ref_idx] or "").strip() if ref_idx < len(r) else ""
                    following = (r[ref_idx + 1] if (ref_idx + 1) < len(r) else "")
                    ref_val = _normalize_avo_ref(raw_ref, following)
                    qty_val = (r[qty_idx] or "").strip() if qty_idx < len(r) else ""
                    if not ref_val or not qty_val: 
                        continue
                    if not material_pat.match(ref_val): 
                        continue
                    qtxt = qty_val.replace("\u00A0", "").replace(" ", "").replace(",", "")
                    if not re.match(r"^-?\d+(?:\.\d+)?$", qtxt):
                        continue
                    _add(ref_val, int(float(qtxt)))
                    found += 1

        # -------- attempt 2: text-line regex fallback --------
        if found == 0:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    if total_row_pat.search(line):
                        continue
                    m = line_pat.search(line)
                    if not m:
                        continue
                    ref_core = m.group(1).strip()
                    ref_sfx = (m.group(2) or "").strip().upper()
                    ref_val = ref_core + (ref_sfx if ref_sfx in SUFFIX_TOKENS else "")
                    qty_val = int(m.group(3))
                    _add(ref_val, qty_val)


    df = pd.DataFrame(rows, columns=["Date","DeliveryNo","AVOMaterialNo","Quantity","Site","Status"])
    if not df.empty:
        df = df.drop_duplicates().reset_index(drop=True)
    return df


@app.route("/preview", methods=["POST"])
def preview():
    file = request.files.get("file")
    file_type = request.form.get("file_type")
    active_tab = 'edi' if file_type == 'EDI' else 'deliveries'

    if not file or not allowed_file(file.filename):
        error_msg = "Invalid file. Please select a .csv, .xlsx, .xls, or .pdf file."
        return render_template_string(
            HTML_PAGE,
            active_tab=active_tab,
            edi_msg=error_msg if active_tab == "edi" else None, edi_ok=False,
            deliv_msg=error_msg if active_tab == "deliveries" else None, deliv_ok=False
        )

    try:
        ext = os.path.splitext(file.filename)[1].lower()
        temp_path = os.path.join(OUTPUT_DIR, f"upload_{int(time.time())}{ext}")
        file.save(temp_path)

        pdf_file_for_embed = None

        if ext in (".xls", ".xlsx"):
            df = pd.read_excel(temp_path)

        elif ext == ".pdf" and file_type == "LIVRAISON":
            # Parse the delivery invoice PDF
            with open(temp_path, "rb") as fh:
                pdf_bytes = fh.read()
            df = parse_delivery_pdf_bytes(pdf_bytes, default_site="Tunisia")
            # enable inline PDF preview in the Delivery tab
            pdf_name = f"pdf_{int(time.time())}.pdf"
            pdf_copy_path = os.path.join(OUTPUT_DIR, secure_filename(pdf_name))
            # keep a copy (you can also just reuse temp_path if you prefer)
            io.open(pdf_copy_path, "wb").write(pdf_bytes)
            pdf_file_for_embed = pdf_name

        else:
            # CSV path
            with open(temp_path, "rb") as f:
                sample = f.read(200_000)
            enc = chardet.detect(sample)["encoding"] or "utf-8"
            df = pd.read_csv(temp_path, encoding=enc, sep=None, engine="python")

        # Persist a temp CSV for insertion reuse
        temp_csv_name = f"preview_{int(time.time())}.csv"
        temp_csv_path = os.path.join(OUTPUT_DIR, temp_csv_name)
        df.to_csv(temp_csv_path, index=False)

        table_html = df.head(20).to_html(index=False, classes="table", table_id="preview-table", border=0)

        # Delivery tab message
        deliv_msg = None
        deliv_ok = None
        edi_msg = None
        edi_ok = None
        if active_tab == "deliveries":
            deliv_msg = f"Parsed {len(df)} rows."
            deliv_ok = True
        else:
            edi_msg = "Preview ready."
            edi_ok = True

        return render_template_string(
            HTML_PAGE,
            table_html=table_html,
            file_type=file_type,
            temp_file=temp_csv_name,
            active_tab=active_tab,
            pdf_file=pdf_file_for_embed,   # << used only when defined
            deliv_msg=deliv_msg, deliv_ok=deliv_ok,
            edi_msg=edi_msg, edi_ok=edi_ok
        )

    except Exception as e:
        error_msg = f"Error processing file: {e}"
        return render_template_string(
            HTML_PAGE,
            active_tab=active_tab,
            edi_msg=error_msg if active_tab == "edi" else None, edi_ok=False,
            deliv_msg=error_msg if active_tab == "deliveries" else None, deliv_ok=False
        )


@app.route("/view/temp/<filename>")
def view_temp_file(filename):
    path = os.path.join(OUTPUT_DIR, secure_filename(filename))
    if not os.path.exists(path):
        abort(404)
    return send_file(path)

def _upsert_sum_delivery(conn, *, site, avo_mat, delivery_no, date, status, qty):
    """
    Add qty to an existing row if (site, avo_mat, delivery_no, date, status) exists,
    otherwise insert a new row. Works without unique constraints, but assumes
    at most one existing row per key (see cleanup note below).
    """
    # Try to add to existing row first
    upd = text("""
        UPDATE "DeliveryDetails"
        SET "Quantity" = "Quantity" + :qty
        WHERE "Site" = :site
          AND "AVOMaterialNo" = :avo_mat
          AND "DeliveryNo" = :del_no
          AND "Date" = :date
          AND "Status" = :status
        RETURNING 1
    """)
    params = {
        "site": site, "avo_mat": avo_mat, "del_no": delivery_no,
        "date": date, "status": status, "qty": int(qty),
    }
    res = conn.execute(upd, params).first()
    if not res:
        ins = text("""
            INSERT INTO "DeliveryDetails"
            ("Site","AVOMaterialNo","DeliveryNo","Quantity","Date","Status")
            VALUES (:site,:avo_mat,:del_no,:qty,:date,:status)
        """)
        conn.execute(ins, params)

SUFFIX_TOKENS = {"PL", "SP"}

def _normalize_avo_ref(s: object, following_hint: object = None) -> str:
    """
    Merge a short type token (PL/SP) into the AVOMaterialNo, removing the space.
    - s: the reference field (may be "V504.243" or "V504.243 PL")
    - following_hint: the next token/cell; if it starts with PL/SP, we merge it.
    """
    base = _safestr(s)
    if not base:
        return ""

    parts = base.split()
    code = parts[0]
    suffix = None

    # case 1: suffix already within the same cell: "V504.243 PL"
    if len(parts) >= 2 and parts[1].upper() in SUFFIX_TOKENS:
        suffix = parts[1].upper()

    # case 2: suffix is in the following cell, e.g. next column starts with "PL" or "SP"
    if not suffix and following_hint is not None:
        nxt = _safestr(following_hint)
        if nxt:
            tok = nxt.split()[0].upper()
            if tok in SUFFIX_TOKENS:
                suffix = tok

    return code + (suffix or "")



def _safestr(v: object) -> str:
    """Return a trimmed string, turning None/NaN into ''. """
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    return str(v).strip()


# ----------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)
